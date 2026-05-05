import argparse
import re
from collections import Counter, defaultdict, deque
from pathlib import Path

from openpyxl import load_workbook


def parse_args():
    parser = argparse.ArgumentParser(description="Match ParentID onto AR item target rows.")
    parser.add_argument("--qne", required=True)
    parser.add_argument("--customer-target", required=True)
    parser.add_argument("--item-target", required=True)
    parser.add_argument("--output", required=True)
    return parser.parse_args()


def norm(value):
    if value is None:
        return None
    text = str(value).replace("_x000D_", " ").replace("_X000D_", " ").replace("\xa0", " ")
    text = " ".join(text.split()).upper()
    return text or None


def loose(value):
    text = norm(value)
    if text is None:
        return None
    return re.sub(r"[^A-Z0-9]+", "", text) or None


def item_key(gl_account, description, project, amount, tax_code):
    return (
        norm(gl_account),
        loose(description),
        norm(project),
        amount,
        norm(tax_code),
    )


def load_customer_parent_map(customer_target_file):
    workbook = load_workbook(customer_target_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    parent_by_doccode = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        parent_id = row[0]
        doc_code = row[1]
        if doc_code not in (None, ""):
            parent_by_doccode[str(doc_code).strip()] = parent_id
    return parent_by_doccode


def build_qne_item_queues(qne_file, parent_by_doccode):
    workbook = load_workbook(qne_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    queues = defaultdict(deque)
    missing_doccodes = Counter()
    current_doc_code = None

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_doc_code = row[0]
        if raw_doc_code not in (None, ""):
            current_doc_code = str(raw_doc_code).strip()

        gl_account = row[43]
        description = row[45]
        tax_code = row[46]
        project = row[49]
        amount = row[54]

        if gl_account in (None, "") and description in (None, "") and amount in (None, ""):
            continue

        doc_code_text = current_doc_code
        parent_id = parent_by_doccode.get(doc_code_text)
        if parent_id is None:
            missing_doccodes[doc_code_text] += 1
            continue

        key = item_key(gl_account, description, project, amount, tax_code)
        queues[key].append(
            {
                "parent_id": parent_id,
                "doc_code": doc_code_text,
                "qne_row_number": row_number,
            }
        )

    return queues, missing_doccodes


def ensure_columns(sheet, header_to_index, columns):
    next_column = sheet.max_column + 1
    for column in columns:
        if column not in header_to_index:
            sheet.cell(row=1, column=next_column, value=column)
            header_to_index[column] = next_column
            next_column += 1


def main():
    args = parse_args()
    parent_by_doccode = load_customer_parent_map(Path(args.customer_target))
    queues, missing_doccodes = build_qne_item_queues(Path(args.qne), parent_by_doccode)

    workbook = load_workbook(Path(args.item_target))
    sheet = workbook[workbook.sheetnames[0]]

    headers = [cell.value for cell in sheet[1]]
    header_to_index = {header: position + 1 for position, header in enumerate(headers)}
    ensure_columns(sheet, header_to_index, ["ParentID", "MatchedDocCode", "MatchedQNERow"])

    summary_sheet_name = "Parent_ID_Match_Summary"
    unmatched_sheet_name = "Parent_ID_Unmatched"
    for sheet_name in [summary_sheet_name, unmatched_sheet_name]:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    matched_rows = 0
    unmatched_rows = []

    for row_number in range(2, sheet.max_row + 1):
        gl_account = sheet.cell(row=row_number, column=header_to_index["GLAccount"]).value
        description = sheet.cell(row=row_number, column=header_to_index["Description"]).value
        project = sheet.cell(row=row_number, column=header_to_index["Project"]).value
        amount = sheet.cell(row=row_number, column=header_to_index["Amount"]).value
        tax_code = sheet.cell(row=row_number, column=header_to_index["TaxCode"]).value

        key = item_key(gl_account, description, project, amount, tax_code)
        match_queue = queues.get(key)

        if match_queue:
            match = match_queue.popleft()
            sheet.cell(row=row_number, column=header_to_index["ParentID"], value=match["parent_id"])
            sheet.cell(row=row_number, column=header_to_index["MatchedDocCode"], value=match["doc_code"])
            sheet.cell(row=row_number, column=header_to_index["MatchedQNERow"], value=match["qne_row_number"])
            matched_rows += 1
        else:
            unmatched_rows.append(
                [
                    row_number,
                    norm(gl_account),
                    description,
                    norm(project),
                    amount,
                    norm(tax_code),
                ]
            )

    leftover_qne_rows = sum(len(queue) for queue in queues.values())

    summary_sheet = workbook.create_sheet(summary_sheet_name)
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["target_item_rows", sheet.max_row - 1])
    summary_sheet.append(["matched_rows", matched_rows])
    summary_sheet.append(["unmatched_rows", len(unmatched_rows)])
    summary_sheet.append(["leftover_qne_item_rows", leftover_qne_rows])
    summary_sheet.append(["missing_customer_doccodes", sum(missing_doccodes.values())])

    unmatched_sheet = workbook.create_sheet(unmatched_sheet_name)
    unmatched_sheet.append(["RowNumber", "GLAccount", "Description", "Project", "Amount", "TaxCode"])
    for row in unmatched_rows:
        unmatched_sheet.append(row)

    workbook.save(Path(args.output))

    print(f"Output written: {args.output}")
    print(f"Matched rows: {matched_rows}")
    print(f"Unmatched rows: {len(unmatched_rows)}")
    print(f"Leftover QNE rows: {leftover_qne_rows}")
    print(f"Missing customer doccodes: {sum(missing_doccodes.values())}")


if __name__ == "__main__":
    main()
