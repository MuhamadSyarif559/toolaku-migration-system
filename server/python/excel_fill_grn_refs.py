import argparse
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path
import re

from openpyxl import load_workbook


def norm(value):
    if value is None:
        return None
    text = " ".join(str(value).replace("\xa0", " ").split()).upper()
    return text or None


def loose_norm(value):
    text = norm(value)
    if text is None:
        return None
    return re.sub(r"[^A-Z0-9]+", "", text) or None


def build_unique_lookup(entries):
    exact = {}
    loose = defaultdict(set)
    for record_id, values in entries:
        for value in values:
            if value in (None, ""):
                continue
            exact[norm(value)] = record_id
            loose[loose_norm(value)].add(record_id)
    return exact, loose


def match_value(value, exact_map, loose_map=None, aliases=None):
    if value in (None, ""):
        return None, "empty"

    key = norm(value)
    if aliases and key in aliases:
        return aliases[key], "alias"

    if key in exact_map:
        return exact_map[key], "exact"

    if loose_map:
        loose_key = loose_norm(value)
        candidates = loose_map.get(loose_key, set())
        if len(candidates) == 1:
            return next(iter(candidates)), "loose"

    return None, "unmatched"


def build_supplier_maps(supplier_file):
    workbook = load_workbook(supplier_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    entries = []
    for supplier_id, supplier_name in sheet.iter_rows(values_only=True, min_row=2):
        entries.append((supplier_id, [supplier_name]))
    exact, loose = build_unique_lookup(entries)
    aliases = {
        norm("BUMIPRIMA CENTURY SDN BHD"): exact.get(norm("BUMI PRIMA CENTURY SDN BHD")),
        norm("VCI TECHNOLOGY SDH BHD"): exact.get(norm("VCI TECHNOLOGY SDN BHD")),
        norm("Q-ART"): exact.get(norm("Q-ART SIGNAGE SDN. BHD")),
    }
    aliases = {key: value for key, value in aliases.items() if value is not None}
    return exact, loose, aliases


def build_term_maps(term_file):
    workbook = load_workbook(term_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    entries = []
    for term_id, code, description in sheet.iter_rows(values_only=True, min_row=2):
        entries.append((term_id, [code, description]))
    exact, loose = build_unique_lookup(entries)
    aliases = {
        norm("C.O.D"): exact.get(norm("C.O.D.")),
    }
    aliases = {key: value for key, value in aliases.items() if value is not None}
    return exact, loose, aliases


def build_location_maps(location_file):
    workbook = load_workbook(location_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    entries = []
    for location_id, code, name in sheet.iter_rows(values_only=True, min_row=2):
        entries.append((location_id, [code, name]))
    exact, loose = build_unique_lookup(entries)
    return exact, loose, {}


def build_currency_maps(currency_file):
    workbook = load_workbook(currency_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    entries = []
    for currency_id, name, symbol in sheet.iter_rows(values_only=True, min_row=2):
        entries.append((currency_id, [name, symbol]))
    exact, loose = build_unique_lookup(entries)
    aliases = {
        norm("RM"): exact.get(norm("MYR")),
        norm("YUAN"): exact.get(norm("CNY")),
    }
    aliases = {key: value for key, value in aliases.items() if value is not None}
    return exact, loose, aliases


def ensure_sheet_removed(workbook, name):
    if name in workbook.sheetnames:
        del workbook[name]


def remove_empty_rows(sheet):
    max_col = sheet.max_column
    max_row = sheet.max_row
    write_row = 2
    removed = 0

    for read_row in range(2, max_row + 1):
        values = [sheet.cell(row=read_row, column=col).value for col in range(1, max_col + 1)]
        if all(value in (None, "") for value in values):
            removed += 1
            continue

        if write_row != read_row:
            for col in range(1, max_col + 1):
                src = sheet.cell(row=read_row, column=col)
                dst = sheet.cell(row=write_row, column=col)
                dst.value = src.value
                if src.has_style:
                    dst._style = copy(src._style)
                else:
                    dst._style = copy(sheet.cell(row=1, column=col)._style)
                if src.number_format:
                    dst.number_format = src.number_format
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.protection = copy(src.protection)
                if src.hyperlink:
                    dst._hyperlink = copy(src.hyperlink)
                else:
                    dst._hyperlink = None
                dst.comment = copy(src.comment) if src.comment else None
            if read_row in sheet.row_dimensions:
                sheet.row_dimensions[write_row].height = sheet.row_dimensions[read_row].height
                sheet.row_dimensions[write_row].hidden = sheet.row_dimensions[read_row].hidden
        write_row += 1

    if removed:
        sheet.delete_rows(write_row, max_row - write_row + 1)

    return removed


def parse_args():
    parser = argparse.ArgumentParser(description="Fill GRN header reference IDs.")
    parser.add_argument("--source", required=True)
    parser.add_argument("--supplier", required=True)
    parser.add_argument("--term", required=True)
    parser.add_argument("--location", required=True)
    parser.add_argument("--currency", required=True)
    parser.add_argument("--output", required=True)
    return parser.parse_args()


def main():
    args = parse_args()
    source_file = Path(args.source)
    supplier_file = Path(args.supplier)
    term_file = Path(args.term)
    location_file = Path(args.location)
    currency_file = Path(args.currency)
    output_file = Path(args.output)

    supplier_exact, supplier_loose, supplier_alias = build_supplier_maps(supplier_file)
    term_exact, term_loose, term_alias = build_term_maps(term_file)
    location_exact, location_loose, location_alias = build_location_maps(location_file)
    currency_exact, currency_loose, currency_alias = build_currency_maps(currency_file)

    workbook = load_workbook(source_file)
    sheet = workbook[workbook.sheetnames[0]]
    removed_empty_rows = remove_empty_rows(sheet)

    headers = [cell.value for cell in sheet[1]]
    index = {name: pos + 1 for pos, name in enumerate(headers)}

    summary_counts = Counter()
    unmatched_rows = []

    for row_number in range(2, sheet.max_row + 1):
        supplier_name = sheet.cell(row=row_number, column=index["CreditorName"]).value
        term_name = sheet.cell(row=row_number, column=index["Term"]).value
        location_name = sheet.cell(row=row_number, column=index["StockLocation"]).value
        currency_name = sheet.cell(row=row_number, column=index["Currency"]).value

        supplier_id, supplier_method = match_value(
            supplier_name, supplier_exact, supplier_loose, supplier_alias
        )
        term_id, term_method = match_value(term_name, term_exact, term_loose, term_alias)
        location_id, location_method = match_value(
            location_name, location_exact, location_loose, location_alias
        )
        currency_id, currency_method = match_value(
            currency_name, currency_exact, currency_loose, currency_alias
        )

        if supplier_id is not None:
            sheet.cell(row=row_number, column=index["Supplierid"], value=supplier_id)
            summary_counts["supplier_matched"] += 1
            summary_counts[f"supplier_{supplier_method}"] += 1
        elif supplier_name not in (None, ""):
            unmatched_rows.append((row_number, "Supplierid", supplier_name))
            summary_counts["supplier_unmatched"] += 1

        if term_id is not None:
            sheet.cell(row=row_number, column=index["Termid"], value=term_id)
            summary_counts["term_matched"] += 1
            summary_counts[f"term_{term_method}"] += 1
        elif term_name not in (None, ""):
            unmatched_rows.append((row_number, "Termid", term_name))
            summary_counts["term_unmatched"] += 1

        if location_id is not None:
            sheet.cell(row=row_number, column=index["Location Id"], value=location_id)
            summary_counts["location_matched"] += 1
            summary_counts[f"location_{location_method}"] += 1
        elif location_name not in (None, ""):
            unmatched_rows.append((row_number, "Location Id", location_name))
            summary_counts["location_unmatched"] += 1

        if currency_id is not None:
            sheet.cell(row=row_number, column=index["Currency id"], value=currency_id)
            summary_counts["currency_matched"] += 1
            summary_counts[f"currency_{currency_method}"] += 1
        elif currency_name not in (None, ""):
            unmatched_rows.append((row_number, "Currency id", currency_name))
            summary_counts["currency_unmatched"] += 1

    ensure_sheet_removed(workbook, "Reference_Match_Summary")
    ensure_sheet_removed(workbook, "Reference_Unmatched")

    summary_sheet = workbook.create_sheet("Reference_Match_Summary")
    summary_sheet.append(["Metric", "Count"])
    summary_sheet.append(["empty_rows_removed", removed_empty_rows])
    for key in sorted(summary_counts):
        summary_sheet.append([key, summary_counts[key]])

    unmatched_sheet = workbook.create_sheet("Reference_Unmatched")
    unmatched_sheet.append(["RowNumber", "TargetColumn", "SourceValue"])
    for row_number, column_name, source_value in unmatched_rows:
        unmatched_sheet.append([row_number, column_name, source_value])

    workbook.save(output_file)

    print(f"Output written: {output_file}")
    print(f"Supplier unmatched: {summary_counts['supplier_unmatched']}")
    print(f"Term unmatched: {summary_counts['term_unmatched']}")
    print(f"Location unmatched: {summary_counts['location_unmatched']}")
    print(f"Currency unmatched: {summary_counts['currency_unmatched']}")


if __name__ == "__main__":
    main()
