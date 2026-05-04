import argparse
from collections import Counter, defaultdict
from pathlib import Path
import re

from openpyxl import load_workbook


FALLBACK_STOCK_ID = 6
FALLBACK_STOCK_NAME = "General Service"


def norm(value):
    if value is None:
        return None
    text = str(value).replace("_x000D_", " ").replace("_X000D_", " ").replace("\xa0", " ")
    text = " ".join(text.split()).upper()
    return text or None


def loose(value):
    text = norm(value)
    if text is None:
        return None
    return re.sub(r"[^A-Z0-9]+", "", text) or None


def looks_code(value):
    text = norm(value)
    return bool(text and re.fullmatch(r"[0-9]+(?:\.[0-9]+)+", text))


def token_set(value):
    text = norm(value)
    if not text:
        return set()
    return {token for token in re.split(r"[^A-Z0-9]+", text) if len(token) >= 4}


def build_stock_refs(stock_file):
    workbook = load_workbook(stock_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    by_name = defaultdict(set)
    by_desc = defaultdict(set)
    by_loose_name = defaultdict(set)
    by_code = defaultdict(set)
    by_prefix = defaultdict(set)
    info = {}

    for stock_id, name, description, stock_code in sheet.iter_rows(values_only=True, min_row=2):
        nname = norm(name)
        ndesc = norm(description)
        ncode = norm(stock_code)
        info[stock_id] = {"name": nname, "desc": ndesc, "code": ncode}

        if nname:
            by_name[nname].add(stock_id)
            by_loose_name[loose(name)].add(stock_id)
        if ndesc and ndesc != "NULL":
            by_desc[ndesc].add(stock_id)
        if ncode:
            by_code[ncode].add(stock_id)
            parts = ncode.split(".")
            if len(parts) >= 3:
                by_prefix[".".join(parts[:3])].add(stock_id)

    return {
        "by_name": by_name,
        "by_desc": by_desc,
        "by_loose_name": by_loose_name,
        "by_code": by_code,
        "by_prefix": by_prefix,
        "info": info,
    }


def build_location_ref(location_file):
    workbook = load_workbook(location_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    mapping = {}
    for location_id, code, name in sheet.iter_rows(values_only=True, min_row=2):
        if code not in (None, ""):
            mapping[norm(code)] = location_id
        if name not in (None, ""):
            mapping[norm(name)] = location_id
    return mapping


def build_tax_ref(tax_file):
    workbook = load_workbook(tax_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    mapping = {}
    for row in sheet.iter_rows(values_only=True):
        if len(row) >= 3 and row[0] not in (None, "") and row[2] not in (None, ""):
            mapping[norm(row[2])] = row[0]
    return mapping


def build_uom_ref(uom_file):
    workbook = load_workbook(uom_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    mapping = defaultdict(list)
    for uom_id, stock_id, einvoice_uom_code, description, code in sheet.iter_rows(
        values_only=True, min_row=2
    ):
        mapping[(stock_id, norm(code))].append(
            {
                "uom_id": uom_id,
                "stock_id": stock_id,
                "einvoice_uom_code": einvoice_uom_code,
                "description": description,
                "code": code,
            }
        )
    return mapping


def learn_stock_code_patterns(sheet, col_index, stock_refs):
    stock_idx = col_index["Stock"]
    desc_idx = col_index["Description"]
    observed = defaultdict(set)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        stock_value = row[stock_idx]
        desc_value = row[desc_idx]
        stock_id, _ = match_stock_from_description(stock_value, desc_value, stock_refs)
        if stock_id is not None and stock_value not in (None, ""):
            observed[norm(stock_value)].add(stock_id)
    return observed


def match_stock_from_description(stock_value, desc_value, stock_refs):
    by_name = stock_refs["by_name"]
    by_desc = stock_refs["by_desc"]
    by_loose_name = stock_refs["by_loose_name"]

    ndesc = norm(desc_value)
    if ndesc in by_name and len(by_name[ndesc]) == 1:
        return next(iter(by_name[ndesc])), "desc_name_exact"
    if ndesc in by_desc and len(by_desc[ndesc]) == 1:
        return next(iter(by_desc[ndesc])), "desc_desc_exact"

    ldesc = loose(desc_value)
    if ldesc in by_loose_name and len(by_loose_name[ldesc]) == 1:
        return next(iter(by_loose_name[ldesc])), "desc_name_loose"

    nstock = norm(stock_value)
    if nstock in by_name and len(by_name[nstock]) == 1:
        return next(iter(by_name[nstock])), "stock_name_exact"

    return None, None


def match_stock(stock_value, desc_value, stock_refs, observed_single):
    stock_id, method = match_stock_from_description(stock_value, desc_value, stock_refs)
    if stock_id is not None:
        return stock_id, method

    nstock = norm(stock_value)
    ndesc = norm(desc_value)
    by_code = stock_refs["by_code"]
    by_prefix = stock_refs["by_prefix"]
    info = stock_refs["info"]

    if nstock:
        learned_ids = observed_single.get(nstock, set())
        if len(learned_ids) == 1:
            return next(iter(learned_ids)), "observed_stock_single"

        if nstock in by_code and len(by_code[nstock]) == 1:
            return next(iter(by_code[nstock])), "stock_code_exact"

        if looks_code(nstock) and len(by_prefix.get(nstock, set())) == 1:
            candidate_id = next(iter(by_prefix[nstock]))
            candidate_tokens = token_set(info[candidate_id]["name"]) | token_set(info[candidate_id]["desc"])
            if token_set(ndesc) & candidate_tokens:
                return candidate_id, "stock_prefix_token"

    if nstock or ndesc:
        return FALLBACK_STOCK_ID, "fallback_general_service"

    return None, "empty"


def match_location(location_value, location_ref):
    nloc = norm(location_value)
    return location_ref.get(nloc)


def match_tax(tax_code_value, tax_ref):
    ntax = norm(tax_code_value)
    return tax_ref.get(ntax)


def uom_aliases(uom_value):
    value = norm(uom_value)
    if value is None:
        return []
    aliases = [value]
    alias_map = {
        "PCS": ["PC", "PCS"],
        "PC(S)": ["PC", "PCS", "PC(S)"],
        "UNIT(S)": ["UNIT(S)", "UNIT"],
        "LITRE ": ["LITRE", "LITER", "L"],
        "LITRE": ["LITRE", "LITER", "L"],
        "LITER": ["LITER", "LITRE", "L"],
        "METER": ["METER", "MTR"],
        "MONTH": ["MONTH", "MTH"],
        "BTL": ["BTL", "BOTTLE"],
        "DR": ["DR", "DRUM"],
        "PKT": ["PKT", "PACKET"],
        "ROLL": ["ROLL", "RL"],
        "PR": ["PR", "PRS"],
        "1/2 BAG": ["1/2 BAG", "BAG"],
        "1/2 DRUM": ["1/2 DRUM", "DRUM"],
    }
    for alias in alias_map.get(value, []):
        if alias not in aliases:
            aliases.append(alias)
    return aliases


def build_uom_preview(stock_id, uom_value, uom_ref):
    if stock_id in (None, "") or uom_value in (None, ""):
        return None, "empty", 0, None

    for alias in uom_aliases(uom_value):
        matches = uom_ref.get((stock_id, alias), [])
        if matches:
            source_uom = norm(uom_value)

            def rank_key(item):
                return (
                    0 if norm(item["code"]) == source_uom else 1,
                    0 if norm(item["description"]) == source_uom else 1,
                    item["uom_id"],
                )

            selected = sorted(matches, key=rank_key)[0]
            note = None
            if len(matches) > 1:
                note = "Multiple UOM rows found for the same StockId + UOM code; selected the best-ranked row"
            return selected, "exact_or_alias", len(matches), note
    return None, "unmatched", 0, None


def ensure_sheet_removed(workbook, name):
    if name in workbook.sheetnames:
        del workbook[name]


def parse_args():
    parser = argparse.ArgumentParser(description="Fill GRN detail reference IDs.")
    parser.add_argument("--source", required=True)
    parser.add_argument("--stock", required=True)
    parser.add_argument("--tax", required=True)
    parser.add_argument("--location", required=True)
    parser.add_argument("--uom", required=True)
    parser.add_argument("--output", required=True)
    return parser.parse_args()


def main():
    args = parse_args()
    source_file = Path(args.source)
    stock_file = Path(args.stock)
    tax_file = Path(args.tax)
    location_file = Path(args.location)
    uom_file = Path(args.uom)
    output_file = Path(args.output)

    stock_refs = build_stock_refs(stock_file)
    location_ref = build_location_ref(location_file)
    tax_ref = build_tax_ref(tax_file)
    uom_ref = build_uom_ref(uom_file)

    workbook = load_workbook(source_file)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    col_index = {name: idx for idx, name in enumerate(headers)}
    col_number = {name: idx + 1 for idx, name in enumerate(headers)}

    observed_single = learn_stock_code_patterns(sheet, col_index, stock_refs)

    summary = Counter()
    fallback_rows = []
    uom_preview_rows = []

    for row_number in range(2, sheet.max_row + 1):
        stock_value = sheet.cell(row=row_number, column=col_number["Stock"]).value
        desc_value = sheet.cell(row=row_number, column=col_number["Description"]).value
        location_value = sheet.cell(row=row_number, column=col_number["StockLocation"]).value
        tax_code_value = sheet.cell(row=row_number, column=col_number["TaxCode"]).value
        uom_value = sheet.cell(row=row_number, column=col_number["UOM"]).value

        stock_id, stock_method = match_stock(stock_value, desc_value, stock_refs, observed_single)
        if stock_id is not None:
            sheet.cell(row=row_number, column=col_number["StockId"], value=stock_id)
            summary["stock_rows_filled"] += 1
            summary[f"stock_{stock_method}"] += 1
            if stock_method == "fallback_general_service":
                fallback_rows.append([row_number, stock_value, desc_value, FALLBACK_STOCK_ID, FALLBACK_STOCK_NAME])

        location_id = match_location(location_value, location_ref)
        if location_id is not None:
            sheet.cell(row=row_number, column=col_number["LocationId"], value=location_id)
            summary["location_rows_filled"] += 1
        elif location_value not in (None, ""):
            summary["location_unmatched"] += 1

        tax_id = match_tax(tax_code_value, tax_ref)
        if tax_id is not None:
            sheet.cell(row=row_number, column=col_number["Taxid"], value=tax_id)
            summary["tax_rows_filled"] += 1
        elif tax_code_value not in (None, ""):
            summary["tax_unmatched"] += 1

        uom_match, uom_method, uom_candidate_count, uom_note = build_uom_preview(
            stock_id, uom_value, uom_ref
        )
        if stock_id is not None or uom_value not in (None, ""):
            if uom_match:
                sheet.cell(row=row_number, column=col_number["UOMID"], value=uom_match["uom_id"])
                summary["uomid_rows_filled"] += 1
                if norm(uom_value) == norm(uom_match["code"]):
                    summary["uomid_exact_code"] += 1
                else:
                    summary["uomid_alias_code"] += 1
                if uom_candidate_count > 1:
                    summary["uomid_ambiguous_selected"] += 1
                uom_preview_rows.append(
                    [
                        row_number,
                        stock_id,
                        stock_value,
                        desc_value,
                        uom_value,
                        uom_match["uom_id"],
                        uom_match["code"],
                        uom_match["einvoice_uom_code"],
                        uom_match["description"],
                        uom_candidate_count,
                        uom_method,
                        uom_note or "Matched and written to UOMID",
                    ]
                )
            elif uom_value not in (None, ""):
                summary["uomid_unmatched"] += 1
                uom_preview_rows.append(
                    [
                        row_number,
                        stock_id,
                        stock_value,
                        desc_value,
                        uom_value,
                        None,
                        None,
                        None,
                        None,
                        0,
                        uom_method,
                        "No stock-specific UOM reference found",
                    ]
                )

    ensure_sheet_removed(workbook, "Match_Summary")
    ensure_sheet_removed(workbook, "Stock_Fallback_Rows")
    ensure_sheet_removed(workbook, "UOM_Reference_Preview")

    summary_sheet = workbook.create_sheet("Match_Summary")
    summary_sheet.append(["Metric", "Count"])
    for key in sorted(summary):
        summary_sheet.append([key, summary[key]])

    fallback_sheet = workbook.create_sheet("Stock_Fallback_Rows")
    fallback_sheet.append(["RowNumber", "SourceStock", "Description", "AssignedStockId", "AssignedStockName"])
    for row in fallback_rows:
        fallback_sheet.append(row)

    uom_sheet = workbook.create_sheet("UOM_Reference_Preview")
    uom_sheet.append(
        [
            "RowNumber",
            "MatchedStockId",
            "SourceStock",
            "Description",
            "SourceUOM",
            "MatchedUOMID",
            "MatchedUOMCode",
            "MatchedEInvoiceUOMCode",
            "MatchedUOMDescription",
            "CandidateCount",
            "MatchMethod",
            "Note",
        ]
    )
    for row in uom_preview_rows:
        uom_sheet.append(row)

    workbook.save(output_file)

    print(f"Output written: {output_file}")
    print(f"Stock rows filled: {summary['stock_rows_filled']}")
    print(f"Fallback rows: {summary['stock_fallback_general_service']}")
    print(f"Location rows filled: {summary['location_rows_filled']}")
    print(f"Tax rows filled: {summary['tax_rows_filled']}")
    print(f"UOMID rows filled: {summary['uomid_rows_filled']}")
    print(f"UOMID unmatched rows: {summary['uomid_unmatched']}")
    print(f"UOM review rows: {len(uom_preview_rows)}")


if __name__ == "__main__":
    main()
