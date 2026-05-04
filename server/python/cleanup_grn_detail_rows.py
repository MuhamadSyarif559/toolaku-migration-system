import argparse
from pathlib import Path

from openpyxl import load_workbook


MAIN_SHEET = "GRN Details Template"
ROW_NUMBER_SHEETS = {"Stock_Fallback_Rows", "UOM_Reference_Preview"}


def is_blank(value):
    return value in (None, "")


def text_value(value):
    if value is None:
        return None
    text = str(value).replace("_x000D_", "\n").strip()
    return text or None


def combine_description(base, addition):
    base_text = text_value(base)
    add_text = text_value(addition)
    if not add_text:
        return base
    if not base_text:
        return add_text
    parts = base_text.splitlines()
    if add_text in parts:
        return base_text
    return f"{base_text}\n{add_text}"


def parse_args():
    parser = argparse.ArgumentParser(description="Clean GRN detail rows.")
    parser.add_argument("--source", required=True)
    parser.add_argument("--output", required=True)
    return parser.parse_args()


def main():
    args = parse_args()
    input_file = Path(args.source)
    output_file = Path(args.output)

    workbook = load_workbook(input_file)
    sheet = workbook[MAIN_SHEET]

    headers = [cell.value for cell in sheet[1]]
    idx = {name: pos for pos, name in enumerate(headers)}

    original_max_row = sheet.max_row
    original_rows = [
        [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 1)]
        for row_num in range(2, sheet.max_row + 1)
    ]

    kept_rows = []
    old_to_new = {}
    pending_notes = []
    current_anchor_index = None

    deleted_rows = 0
    merged_note_rows = 0
    merged_note_texts = 0

    for old_row_number, row in enumerate(original_rows, start=2):
        qty = row[idx["Qty"]]
        uom = row[idx["UOM"]]
        location_id = row[idx["LocationId"]]
        stock_location = row[idx["StockLocation"]]
        description = row[idx["Description"]]

        has_location = not is_blank(location_id) or not is_blank(stock_location)
        has_detail_uom_qty = not is_blank(uom) and qty not in (None, "", 0)
        should_delete = is_blank(uom) and qty in (None, "", 0) and not has_location

        if has_location:
            pending_notes = []
            current_anchor_index = None

        if should_delete:
            deleted_rows += 1
            note_text = text_value(description)
            if note_text:
                if current_anchor_index is not None:
                    anchor_row = kept_rows[current_anchor_index]
                    anchor_row[idx["Description"]] = combine_description(
                        anchor_row[idx["Description"]], note_text
                    )
                    merged_note_rows += 1
                    merged_note_texts += 1
                else:
                    pending_notes.append(note_text)
            continue

        new_row = list(row)
        if has_detail_uom_qty and pending_notes:
            for note in pending_notes:
                new_row[idx["Description"]] = combine_description(new_row[idx["Description"]], note)
                merged_note_texts += 1
            pending_notes = []

        kept_rows.append(new_row)
        old_to_new[old_row_number] = len(kept_rows) + 1

        if has_detail_uom_qty:
            current_anchor_index = len(kept_rows) - 1

    if pending_notes and current_anchor_index is not None:
        anchor_row = kept_rows[current_anchor_index]
        for note in pending_notes:
            anchor_row[idx["Description"]] = combine_description(anchor_row[idx["Description"]], note)
            merged_note_texts += 1
            merged_note_rows += 1

    for new_row_number, row_values in enumerate(kept_rows, start=2):
        for col_number, value in enumerate(row_values, start=1):
            sheet.cell(row=new_row_number, column=col_number, value=value)

    final_data_rows = len(kept_rows)
    if final_data_rows + 1 < original_max_row:
        sheet.delete_rows(final_data_rows + 2, original_max_row - final_data_rows - 1)

    for sheet_name in ROW_NUMBER_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        ref_sheet = workbook[sheet_name]
        ref_headers = [cell.value for cell in ref_sheet[1]]
        ref_rows = [
            [ref_sheet.cell(row=row_num, column=col_num).value for col_num in range(1, ref_sheet.max_column + 1)]
            for row_num in range(2, ref_sheet.max_row + 1)
        ]
        kept_ref_rows = []
        for row in ref_rows:
            source_row_number = row[0]
            if source_row_number in old_to_new:
                updated = list(row)
                updated[0] = old_to_new[source_row_number]
                kept_ref_rows.append(updated)

        for new_row_number, row_values in enumerate(kept_ref_rows, start=2):
            for col_number, value in enumerate(row_values, start=1):
                ref_sheet.cell(row=new_row_number, column=col_number, value=value)

        if len(kept_ref_rows) + 1 < ref_sheet.max_row:
            ref_sheet.delete_rows(len(kept_ref_rows) + 2, ref_sheet.max_row - len(kept_ref_rows) - 1)

        for col_number, header in enumerate(ref_headers, start=1):
            ref_sheet.cell(row=1, column=col_number, value=header)

    if "Cleanup_Summary" in workbook.sheetnames:
        del workbook["Cleanup_Summary"]
    summary = workbook.create_sheet("Cleanup_Summary")
    summary.append(["Metric", "Count"])
    summary.append(["original_data_rows", len(original_rows)])
    summary.append(["final_data_rows", final_data_rows])
    summary.append(["deleted_rows", deleted_rows])
    summary.append(["merged_note_rows", merged_note_rows])
    summary.append(["merged_note_texts", merged_note_texts])
    summary.append(
        [
            "rows_kept_with_location",
            sum(
                1
                for row in kept_rows
                if not is_blank(row[idx["LocationId"]]) or not is_blank(row[idx["StockLocation"]])
            ),
        ]
    )

    workbook.save(output_file)

    print(f"Output written: {output_file}")
    print(f"Original data rows: {len(original_rows)}")
    print(f"Final data rows: {final_data_rows}")
    print(f"Deleted rows: {deleted_rows}")
    print(f"Merged note texts: {merged_note_texts}")


if __name__ == "__main__":
    main()
